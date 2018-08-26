import pandas as pd
from scipy import stats
import re
import os
from openpyxl.reader.excel import load_workbook
from pandas import DataFrame
from bs4 import BeautifulSoup
from nltk import tokenize
from keras.preprocessing.text import Tokenizer, text_to_word_sequence
from keras.preprocessing.sequence import pad_sequences
from keras.utils.np_utils import to_categorical
from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer
from model_hatt import *
from model_textCNN import *
import matplotlib.pyplot as plt

xlsx_file = './data/MAUDE_2008_2016_review.xlsx'
tsv_file = './data/MAUDE_2008_2016_review.tsv'


def xlsx_to_tsv():
    """
    Transform xlsx file to tsv file
    """
    df_list = []
    sum_pos = 0
    sum_neg = 0

    wb = load_workbook(xlsx_file)
    print('====== DATA SUMMARY ======')
    print('.xlsx file from:', xlsx_file)
    print('YEAR', 'HIT', 'Non-H', 'ALL')

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        df = DataFrame(ws.values)
        pos = 0
        neg = 0
        for index, row in df.iterrows():
            if row[2] == 1:
                pos += 1
            else:
                neg += 1
        sum_pos += pos
        sum_neg += neg
        df_list.append(df)
        print(sheetname, pos, neg, pos + neg)

    wb.close()
    print('SUM', sum_pos, sum_neg, sum_pos + sum_neg)

    with open(tsv_file, "w") as f:
        f.write("ID\tHIT\tREPORT\n")
        for one_year in df_list:
            for index, row in one_year.iterrows():
                f.write(str(row[0]) + "\t" + str(row[2]) + "\t" + str(row[1]) + "\n")

    print('.tsv file written to:', tsv_file)


def clean_str(string):
    """
    Tokenization/string cleaning for dataset
    Every dataset is lower cased except
    """
    string = re.sub(r"\\", "", str(string))
    string = re.sub(r"\'", "", string)
    string = re.sub(r"\"", "", string)
    return string.strip().lower()


def read_MAUDE():
    """
    Read MAUDE data from .tsv file

    Return:
    documents_sent -- List[document #][sentence #], tokenized sentences
    labels -- List[document #], HIT labels
    documents -- List[document #], list of individual reports
    """
    print("Reading raw data ...")
    data_train = pd.read_csv(tsv_file, sep='\t')
    print("Raw data shape: " + str(data_train.shape))

    documents_sent = []
    labels = []
    documents = []

    print("Processing sentence tokenization ...")
    for idx in range(data_train.REPORT.shape[0]):
        text = BeautifulSoup(data_train.REPORT[idx], "html5lib")
        text = clean_str(text.get_text().encode('ascii', 'ignore'))
        documents.append(text)
        sentences = tokenize.sent_tokenize(text)
        documents_sent.append(sentences)
        labels.append(data_train.HIT[idx])

    sent_in_doc = []
    for doc in documents_sent:
        sent_in_doc.append(len(doc))
    print('Max sentences # in a doc:', np.max(sent_in_doc))
    print('Min sentences # in a doc:', np.min(sent_in_doc))
    print('Mean sentences # in a doc:', np.mean(sent_in_doc))
    print('Median sentences # in a doc:', np.median(sent_in_doc))
    interval = stats.norm.interval(0.95, np.mean(sent_in_doc), np.std(sent_in_doc))
    print('95% confidence interval of sentences # in a doc:', interval)

    return documents_sent, labels, documents


def read_MAUDE_by_document():
    """
    Read MAUDE data from .tsv file

    Return:
    labels -- List[document #], HIT labels
    documents -- List[document #], list of individual reports
    """
    print("Reading raw data ...")
    data_train = pd.read_csv(tsv_file, sep='\t')
    print("Raw data shape: " + str(data_train.shape))

    labels = []
    documents = []

    print("Processing sentence tokenization ...")
    for idx in range(data_train.REPORT.shape[0]):
        text = BeautifulSoup(data_train.REPORT[idx], "html5lib")
        text = clean_str(text.get_text().encode('ascii', 'ignore'))
        documents.append(text)
        labels.append(data_train.HIT[idx])

    word_in_doc = []
    for doc in documents:
        word_in_doc.append(doc.count(' ') + 1)
    print('Max word # in a doc:', np.max(word_in_doc))
    print('Min word # in a doc:', np.min(word_in_doc))
    print('Mean word # in a doc:', np.mean(word_in_doc))
    print('Median word # in a doc:', np.median(word_in_doc))
    interval = stats.norm.interval(0.95, np.mean(word_in_doc), np.std(word_in_doc))
    print('95% confidence interval of word # in a doc:', interval)

    return labels, documents


def read_MAUDE_simple():
    """
        Read MAUDE data from .tsv file

        Return:
        labels -- List[document #], HIT labels
        documents -- List[document #], list of individual reports
    """
    print("Reading raw data ...")
    data_train = pd.read_csv(tsv_file, sep='\t')
    print("Raw data shape: " + str(data_train.shape))

    labels = []
    documents = []

    for idx in range(data_train.REPORT.shape[0]):
        documents.append(data_train.REPORT[idx])
        labels.append(data_train.HIT[idx])

    return labels, documents


def read_MAUDE_hierarchical_simple():
    """
        Read MAUDE data from .tsv file

        Return:
        documents_sent -- List[document #][sentence #], tokenized sentences
        labels -- List[document #], HIT labels
        documents -- List[document #], list of individual reports
    """
    print("Reading raw data ...")
    data_train = pd.read_csv(tsv_file, sep='\t')
    print("Raw data shape: " + str(data_train.shape))

    documents_sent = []
    labels = []
    documents = []

    for idx in range(data_train.REPORT.shape[0]):
        documents.append(data_train.REPORT[idx])
        sentences = tokenize.sent_tokenize(data_train.REPORT[idx])
        documents_sent.append(sentences)
        labels.append(data_train.HIT[idx])

    sent_in_doc = []
    word_in_doc = []
    word_in_sent = []

    for doc in documents_sent:
        sent_in_doc.append(len(doc))
        for sent in doc:
            word_in_sent.append(sent.count(' ') + 1)
    print('-- sentences # in a doc --')
    print('Max:', np.max(sent_in_doc))
    print('Min:', np.min(sent_in_doc))
    print('Mean:', np.mean(sent_in_doc))
    print('Median:', np.median(sent_in_doc))
    interval = stats.norm.interval(0.95, np.mean(sent_in_doc), np.std(sent_in_doc))
    print('95% confidence interval:', interval)

    print('-- word # in a sentence --')
    print('Max:', np.max(word_in_sent))
    print('Min:', np.min(word_in_sent))
    print('Mean:', np.mean(word_in_sent))
    print('Median:', np.median(word_in_sent))
    interval = stats.norm.interval(0.95, np.mean(word_in_sent), np.std(word_in_sent))
    print('95% confidence interval:', interval)

    for doc in documents:
        word_in_doc.append(doc.count(' ') + 1)
    print('-- word # in a doc --')
    print('Max:', np.max(word_in_doc))
    print('Min:', np.min(word_in_doc))
    print('Mean:', np.mean(word_in_doc))
    print('Median:', np.median(word_in_doc))
    interval = stats.norm.interval(0.95, np.mean(word_in_doc), np.std(word_in_doc))
    print('95% confidence interval:', interval)

    return documents_sent, labels, documents


def read_MAUDE_train_dev_test(train_dev_test):
    """
        Read MAUDE data from .tsv file

        Arguments:
        train_dev_test -- file of train, dev, or test

        Return:
        documents_sent -- List[document #][sentence #], tokenized sentences
        labels -- List[document #], HIT labels
        documents -- List[document #], list of individual reports
    """
    data_train = pd.read_csv(train_dev_test, sep='\t')

    documents_sent = []
    labels = []
    documents = []

    for idx in range(data_train.REPORT.shape[0]):
        documents.append(data_train.REPORT[idx])
        sentences = tokenize.sent_tokenize(data_train.REPORT[idx])
        documents_sent.append(sentences)
        labels.append(data_train.HIT[idx])

    return documents_sent, to_categorical(np.asarray(labels)), np.asarray(documents)


def tokenazation(documents_sent, labels, documents, MAX_NUM_WORDS, MAX_SENTS, MAX_SENT_LENGTH):
    """
    Tokenization

    Arguments:
    documents_sent -- List[document #][sentence #], tokenized sentences
    labels -- List[document #], HIT labels
    documents -- List[document #], list of individual reports
    MAX_NUM_WORDS -- Positive integer, the maximum number of words to keep in a document, based on word frequency
    MAX_SENTS -- Positive integer, maximum sentence # in a document
    MAX_SENT_LENGTH -- Positive integer, maximum word # in a sentence

    Return:
    data -- shape:(document #, sentence #, word #), word_index after tokenization
    labels -- shape:(document #, label #)
    word_index -- dict, key: tokens ranked by frequence, value: ranking index
    """
    print("Fitting text by MAX_NUM_WORDS = " + str(MAX_NUM_WORDS) + " ...")
    tokenizer = Tokenizer(num_words=MAX_NUM_WORDS)
    tokenizer.fit_on_texts(documents)

    data = np.zeros((len(documents), MAX_SENTS, MAX_SENT_LENGTH), dtype='int32')

    print("Putting word token indexes into input matrix ...")
    for i, sentences in enumerate(documents_sent):
        for j, sent in enumerate(sentences):
            if j < MAX_SENTS:
                wordTokens = text_to_word_sequence(sent)
                k = 0
                for _, word in enumerate(wordTokens):
                    if k < MAX_SENT_LENGTH and tokenizer.word_index[word] < MAX_NUM_WORDS:
                        data[i, j, k] = tokenizer.word_index[word]
                        k = k + 1

    word_index = tokenizer.word_index
    print("Total unique tokens: " + str(len(word_index)))

    labels = to_categorical(np.asarray(labels))
    print('Shape of data tensor:', data.shape)
    print('Shape of label tensor:', labels.shape)

    return data, labels, word_index


def tokenazation_combined_models(word_index, documents_sent, MAX_NUM_WORDS, MAX_SENTS, MAX_SENT_LENGTH):
    """
    Tokenization for combined models

    Arguments:
    word_index -- tokenizer fitted by training data
    documents_sent -- List[document #][sentence #], tokenized sentences
    MAX_NUM_WORDS -- Positive integer, the maximum number of words to keep in a document, based on word frequency
    MAX_SENTS -- Positive integer, maximum sentence # in a document
    MAX_SENT_LENGTH -- Positive integer, maximum word # in a sentence

    Return:
    data -- shape:(document #, sentence #, word #), word_index after tokenization
    """
    data = np.zeros((len(documents_sent), MAX_SENTS, MAX_SENT_LENGTH), dtype='int32')
    for i, sentences in enumerate(documents_sent):
        for j, sent in enumerate(sentences):
            if j < MAX_SENTS:
                wordTokens = text_to_word_sequence(sent)
                k = 0
                for _, word in enumerate(wordTokens):
                    if k < MAX_SENT_LENGTH and word_index[word] < MAX_NUM_WORDS:
                        data[i, j, k] = word_index[word]
                        k = k + 1

    return data


def tokenazation_CNN(labels, documents, MAX_NUM_WORDS, MAX_SEQUENCE_LENGTH):
    """
    Tokenization

    Arguments:
    labels -- List[document #], HIT labels
    documents -- List[document #], list of individual reports
    MAX_NUM_WORDS -- Positive integer, the maximum number of words to keep in a document, based on word frequency
    MAX_SEQUENCE_LENGTH -- Positive integer, maximum sequence #

    Return:
    data -- shape:(document #, word #), word_index after tokenization
    labels -- shape:(document #, label #)
    word_index -- dict, key: tokens ranked by frequence, value: ranking index
    """
    print("Fitting text by MAX_NUM_WORDS = " + str(MAX_NUM_WORDS) + " ...")
    tokenizer = Tokenizer(num_words=MAX_NUM_WORDS)
    tokenizer.fit_on_texts(documents)
    sequences = tokenizer.texts_to_sequences(documents)
    word_index = tokenizer.word_index
    print("Total unique tokens: " + str(len(word_index)))

    data = pad_sequences(sequences, maxlen=MAX_SEQUENCE_LENGTH)

    labels = to_categorical(np.asarray(labels))
    print('Shape of data tensor:', data.shape)
    print('Shape of label tensor:', labels.shape)

    return data, labels, word_index


def tokenazation_tfidf(x_train_tfidf, x_test_tfidf):
    """
    Tokenization using tfidf

    Arguments:
    x_train_tfidf -- shape:(document #), x train tfidf
    x_test_tfidf -- shape:(document #), x test tfidf

    Return:
    x_train_tfidf -- shape:(document #, word #), x train tfidf
    x_test_tfidf -- shape:(document #, word), x test tfidf
    """
    count_vec = CountVectorizer()
    x_train_counts = count_vec.fit_transform(x_train_tfidf)
    x_test_counts = count_vec.transform(x_test_tfidf)

    tfidf_tran = TfidfTransformer()
    x_train_tfidf = tfidf_tran.fit_transform(x_train_counts)
    x_test_tfidf = tfidf_tran.transform(x_test_counts)

    x_train_tfidf = x_train_tfidf.todense()
    x_test_tfidf = x_test_tfidf.todense()

    return x_train_tfidf, x_test_tfidf


def prepare_train_dev(data, labels, DATA_SPLIT):
    """
    Preparing training and developing data

    Arguments:
    data -- shape:(document #, sentence #, word #), word_index after tokenization
    labels -- shape:(document #, label #)
    DATA_SPLIT -- digit array with length 2 or 3, with sum of 10. Ratio of train, dev and test data

    Return:
    x_train -- shape:(document #, sentence #, word #), x train
    y_train -- shape:(document #, label #), y train
    x_val -- shape:(document #, sentence #, word #), x dev
    y_val -- shape:(document #, label #), y dev
    x_test -- shape:(document #, sentence #, word #), x test
    y_test -- shape:(document #, label #), y test
    """
    indices = np.arange(data.shape[0])
    np.random.shuffle(indices)
    data = data[indices]
    labels = labels[indices]
    print("Dataset Shuffled.")

    nb_train_samples = int(0.1 * DATA_SPLIT[0] * data.shape[0])
    nb_val_samples = int(0.1 * DATA_SPLIT[1] * data.shape[0])
    x_train = data[: nb_train_samples]
    y_train = labels[: nb_train_samples]
    x_val = data[nb_train_samples: (nb_train_samples + nb_val_samples)]
    y_val = labels[nb_train_samples: (nb_train_samples + nb_val_samples)]
    x_test = data[(nb_train_samples + nb_val_samples):]
    y_test = labels[(nb_train_samples + nb_val_samples):]

    print(str(len(x_train)) + " (0." + str(DATA_SPLIT[0]) + ") for training, " +
          str(len(x_val)) + " (0." + str(DATA_SPLIT[1]) + ") for validation, " +
          str(len(x_test)) + " (0." + str(DATA_SPLIT[2]) + ") for testing.")

    print('Number of positive and negative reviews in training and validation set: ')
    print('Train:', y_train.sum(axis=0))
    print('Val:', y_val.sum(axis=0))
    print('Test:', y_test.sum(axis=0))

    return x_train, y_train, x_val, y_val, x_test, y_test


def prepare_train_dev_baseline(data_we, documents, labels, DATA_SPLIT):
    """
    Preparing training and developing data

    Arguments:
    data_we -- shape:(document #, word #), word_index after tokenization
    documents -- shape:(document #), raw documents
    labels -- shape:(document #, label #)
    DATA_SPLIT -- digit array with length 2 or 3, with sum of 10. Ratio of train, dev and test data

    Return:
    x_train_we -- shape:(document #, word embedding dimension #), x train we
    x_train_tfidf -- shape:(document #), x train tfidf
    y_train -- shape:(document #, label #), y train
    x_test_we -- shape:(document #, word embedding dimension #), x test
    x_test_tfidf -- shape:(document #), x test tfidf
    y_test -- shape:(document #, label #), y test
    """
    indices = np.arange(data_we.shape[0])
    np.random.shuffle(indices)
    data_we = data_we[indices]
    data_tfidf = documents[indices]
    labels = labels[indices]
    print("Dataset Shuffled.")

    nb_train_samples = int(0.1 * DATA_SPLIT[0] * data_we.shape[0])
    nb_val_samples = int(0.1 * DATA_SPLIT[1] * data_we.shape[0])
    x_train_we = data_we[: nb_train_samples]
    x_train_tfidf = data_tfidf[: nb_train_samples]
    y_train = labels[: nb_train_samples]

    x_test_we = data_we[(nb_train_samples + nb_val_samples):]
    x_test_tfidf = data_tfidf[(nb_train_samples + nb_val_samples):]
    y_test = labels[(nb_train_samples + nb_val_samples):]

    print(str(len(x_train_we)) + " (0." + str(DATA_SPLIT[0]) + ") for training, " +
          str(len(x_test_we)) + " (0." + str(DATA_SPLIT[2]) + ") for testing.")

    print('Number of positive and negative reviews in training and validation set: ')
    print('Train:', y_train.sum(axis=0))
    print('Test:', y_test.sum(axis=0))

    return x_train_we, x_train_tfidf, y_train, x_test_we, x_test_tfidf, y_test


def calculate_embedding_matrix(EMB_DIM, word_index):
    """
    Generate word embedding matrix for embedding layer establishment

    Arguments:
    GLOVE_DIR -- String, GloVe file direction
    EMB_DIM -- Positive integer, word embedding dimension, i.e., 50/100/200/300
    word_index -- dict, key: tokens ranked by frequence, value: ranking index

    Return:
    embedding_matrix -- shape:((len(word_index) + 1, EMB_DIM), embedding matrix for building embedding layer
    """
    embeddings_index = {}
    f = open(os.path.join('./data/glove.6B.' + str(EMB_DIM) + 'd.txt'))
    for line in f:
        values = line.split()
        word = values[0]
        coefs = np.asarray(values[1:], dtype='float32')
        embeddings_index[word] = coefs
    f.close()

    print("Total tokens in word embedding resource: " + str(len(embeddings_index)))
    print("Dimensions of word embedding: " + str(EMB_DIM))

    embedding_matrix = np.random.random((len(word_index) + 1, EMB_DIM))
    for word, i in word_index.items():
        embedding_vector = embeddings_index.get(word)
        if embedding_vector is not None:
            # words not found in embedding index will be all-zeros.
            embedding_matrix[i] = embedding_vector

    return embedding_matrix


def mean_emb_for_document(x_word_index, embedding_matrix):
    """
    Calculate mean word embedding for each document by averaging embeddings among all the words

    Arguments:
    x_word_index -- shape:(document #, sentence #, word #), word_index after tokenization
    embedding_matrix -- shape:((len(word_index) + 1, EMB_DIM), embedding matrix for building embedding layer

    Return:
    mean_emb --  shape:(document #, embedding dim #), averaged embedding of each document
    """
    mean_emb = np.zeros((x_word_index.shape[0], embedding_matrix.shape[1]))
    for i in range(x_word_index.shape[0]):
        word_num_in_doc = 0
        for j in range(x_word_index.shape[1]):
            for k in range(x_word_index.shape[2]):
                if x_word_index[i][j][k] != 0:
                    mean_emb[i] += embedding_matrix[x_word_index[i][j][k]]
                    word_num_in_doc += 1
        if word_num_in_doc != 0:
            mean_emb[i] /= word_num_in_doc
    return mean_emb


def mean_emb(x_word_index, embedding_matrix):
    """
    Calculate mean word embedding for each document by averaging embeddings among all the words

    Arguments:
    x_word_index -- shape:(document #, word #), word_index after tokenization
    embedding_matrix -- shape:((len(word_index) + 1, EMB_DIM), embedding matrix for building embedding layer

    Return:
    mean_emb --  shape:(document #, embedding dim #), averaged embedding of each document
    """
    mean_emb = np.zeros((x_word_index.shape[0], embedding_matrix.shape[1]))
    for i in range(x_word_index.shape[0]):
        word_num = 0
        for j in range(x_word_index.shape[1]):
            if x_word_index[i][j] != 0:
                mean_emb[i] += embedding_matrix[x_word_index[i][j]]
                word_num += 1
        if word_num != 0:
            mean_emb[i] /= word_num
    return mean_emb
